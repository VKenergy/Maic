from flask import Flask, render_template, request, send_file
import openpyxl
import requests
from datetime import datetime, timezone, timedelta
import pytz
from io import BytesIO

app = Flask(__name__)

# Список пристроїв (назва та ID)
devices = [
    ("ЩС 4ч", 1828729450),
    ("Щитова СБХМ", 1828726747),
    ("ШС 4к", 1828729467),
    ("ШР-1 (6034)", 1828726749),
    ("ШР КІПА", 1828724649),
    ("ШР 6к чердак", 1828729448),
    ("Чіллер 3-й поверх", 1828729073),
    ("СБХМ-2", 1828726667),
    ("СБХМ-1", 1828726736),
    ("сбхм", 1828724853),
    ("РП-45", 1828729481),
    ("РП-31", 1828729035),
    ("РП-27", 1828729470),
    ("РП-16", 1828726738),
    ("РП-15", 1828724652),
    ("РП Глазурування 3-й поверх", 1828729067),
    ("РП - 41,42", 1828729478),
    ("РП - 39,40", 1828729052),
    ("РП - 36,38", 1828729469),
    ("РП 31 + РП 10", 1828729462),
    ("РП 10.1", 1828729449),
    ("РП 6-2", 1828729040),
    ("РП 5к", 1828729071),
    ("РП 5-5", 1828729484),
    ("РП 5-3", 1828729046),
    ("РП 5-2", 1828729051),
    ("РП 5-1", 1828729474),
    ("РП 5-0", 1828729030),
    ("РП 4.5", 1828729454),
    ("РП 4.3", 1828729471),
    ("РП 4.2", 1828729477),
    ("РП 4-4", 1828729480),
    ("РП 4-1", 1828729452),
    ("РП 4 + РП 6", 1828729457),
    ("РП 3-3", 1828729054),
    ("РП 3-2", 1828726737),
    ("РП 3-1", 1828724862),
    ("Прачка", 1828724651),
    ("ПР-27", 1828726750),
    ("КВ", 1828726735),
    ("Ввід №4 БЦ", 1828726742),
    ("Ввід №2 БЦ", 1828726753),
    ("QF30 Патока", 1828729483)
]

API_KEY = "2154049356"
API_URL = "https://dash.smart-maic.com/api"
kyiv_tz = pytz.timezone("Europe/Kyiv")

def date_to_unix(date):
    return int(date.timestamp())

def fetch_device_data(device_id, date):
    date_7am = date.replace(hour=7, minute=0, second=0, microsecond=0)
    date_7am_utc = date_7am.astimezone(timezone.utc)
    api_url = f"{API_URL}?devid={device_id}&date1={date_to_unix(date_7am_utc)}&date2={date_to_unix(date_7am_utc + timedelta(hours=1))}&period=hour&apikey={API_KEY}"
    response = requests.get(api_url)
    if response.status_code == 200:
        data = response.json()
        if data:
            return data[0].get("Wh1", "Н/Д"), data[0].get("Wh2", "Н/Д"), data[0].get("Wh3", "Н/Д")
    return "Н/Д", "Н/Д", "Н/Д"

def create_excel(start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Дані"

    # Запис заголовків
    ws['A1'] = "Дата і час"
    ws['A2'] = "ID"
    for col, (name, device_id) in enumerate(devices, start=2):
        ws.cell(row=1, column=col, value=name)
        ws.cell(row=2, column=col, value=device_id)

    current_date = start_date
    row = 3
    while current_date <= end_date:
        date_7am = current_date.replace(hour=7, minute=0, second=0, microsecond=0)
        ws.cell(row=row, column=1, value=date_7am.strftime('%Y-%m-%d %H:%M'))
        for col, (name, device_id) in enumerate(devices, start=2):
            Wh1, Wh2, Wh3 = fetch_device_data(device_id, current_date)
            try:
                Wh1 = float(Wh1)
                Wh2 = float(Wh2)
                Wh3 = float(Wh3)
                SUM_kWh = round(((Wh1 + Wh2 + Wh3) / 3) / 1000)
            except ValueError:
                SUM_kWh = "Н/Д"
            ws.cell(row=row, column=col, value=SUM_kWh)
        current_date += timedelta(days=1)
        row += 1

    # Встановлення ширини стовпців
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    # Збереження у BytesIO
    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)
    return excel_data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/data', methods=['POST'])
def data():
    start_date = request.form['start_date']
    end_date = request.form['end_date']
    start_date = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=kyiv_tz)
    end_date = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=kyiv_tz)
    excel_data = create_excel(start_date, end_date)
    return send_file(excel_data, download_name='Maic_data.xlsx', as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
